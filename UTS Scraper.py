import requests
import openpyxl
from bs4 import BeautifulSoup
import re

url = "https://www.handbook.uts.edu.au/"

def getDegrees(): #Get all degrees available on the UTS handbook
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOK.xlsx')
    sheet = wb['Degrees']
    courses = ["bus", "comm", "cii", "dab", "edu", "eng", "health", "health-gem", "it", "intl", "law", "sci", "tdi"]
    graduatetype = ["ug", "pg"]
    degrees = 1
    for i in courses:
        soup = BeautifulSoup(requests.get(url+"/"+i+"/"+graduatetype[0]).content, 'html.parser')  
        tables = soup.find(class_='ie-images').find_all('table')
        for row in tables:
            cols = row.find_all('td')
            cols = [ele.text.strip() for ele in cols]
            for j in cols:
                sheet.cell(degrees, column=1).value = i
                if degrees%2==1:
                    sheet.cell(degrees, column=2).value = j
                else:
                    sheet.cell(degrees, column=3).value = j
                degrees += 1
    wb.save('UTSHANDBOOK.xlsx')
    

#Courses with no structure (just standardised sturcture e.g no core, major or submajor choice will be blank)
def getMajors(): #Get all majors for each degree
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    sheet1 = wb['Degrees']
    sheet2 = wb['Degree MAJ, SUB, CORE']
    degrees = sheet1['B']
    count = 2
    sheet2.cell(1, column=1).value = "Degree Code"
    sheet2.cell(1, column=2).value = "Major Codes"
    sheet2.cell(1, column=3).value = "SubMajor Codes"
    sheet2.cell(1, column=4).value = "Core Codes"
    for i in degrees:
        print(i)
        majors = []
        submajors = []
        core = []
        soup = BeautifulSoup(requests.get(url+"courses/"+i.value.lower()).content, 'html.parser')  
        tables = soup.find_all('table')
        for j in tables:
            list_of_majors = j.find_all('a')
            for k in list_of_majors:
                if "MAJ" in str(k):
                    majors.append(k.text.strip())
                if "SMJ" in str(k):
                    submajors.append(k.text.strip())
                if "STM" in str(k):
                    core.append(k.text.strip())
        sheet2.cell(count, column=1).value = i.value
        sheet2.cell(count, column=2).value = str(majors)
        sheet2.cell(count, column=3).value = str(submajors)
        sheet2.cell(count, column=4).value = str(core)
        count+= 1
    wb.save('UTSHANDBOOKtest.xlsx')

def getCoreSubjects():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    sheet1 = wb['Degree MAJ, SUB, CORE']
    sheet2 = wb['Core Subjects']
    degrees = sheet1['D']
    sheet2.cell(1, column=1).value = "Degree Code"
    sheet2.cell(1, column=2).value = "Core Subjects"
    count = 2
    for i in degrees:
        core_subjects = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        core_codes = [x for x in values1 if x]
        print(core_codes)
        for j in core_codes:
            soup = BeautifulSoup(requests.get(url+"directory/"+j.lower()).content, 'html.parser')
            tables = soup.find_all('table')
            for k in tables:
                a = k.find_all('a')
                for l in a:
                    result = re.search('>(.*)<', str(l))
                    core_subjects.append(result.group(1))
        sheet2.cell(count, column=1).value = i.value
        sheet2.cell(count, column=2).value = str(core_subjects)
        count +=1

    wb.save('UTSHANDBOOKtest.xlsx')

def getMajorSubjects():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    sheet1 = wb['Degree MAJ, SUB, CORE']
    sheet2 = wb['Major Subjects']
    degrees = sheet1['B']
    sheet2.cell(1, column=1).value = "Degree Code"
    sheet2.cell(1, column=2).value = "Major Subjects"
    count = 2
    for i in degrees:
        core_subjects = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        major_codes = [x for x in values1 if x]
        print(major_codes)
        for j in major_codes:
            soup = BeautifulSoup(requests.get(url+"directory/"+j.lower()).content, 'html.parser')
            tables = soup.find_all('table')
            for k in tables:
                a = k.find_all('a')
                for l in a:
                    result = re.search('>(.*)<', str(l))
                    core_subjects.append(result.group(1))
        sheet2.cell(count, column=1).value = i.value
        sheet2.cell(count, column=2).value = str(core_subjects)
        count +=1
    wb.save('UTSHANDBOOKtest.xlsx')

def getSubMajorSubjects():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    sheet1 = wb['Degree MAJ, SUB, CORE']
    sheet2 = wb['SubMajor Subjects']
    degrees = sheet1['C']
    sheet2.cell(1, column=1).value = "Degree Code"
    sheet2.cell(1, column=2).value = "Sub Major Subjects"
    count = 2
    for i in degrees:
        sub_subjects = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        sub_major_codes = [x for x in values1 if x]
        print(sub_major_codes)
        for j in sub_major_codes:
            soup = BeautifulSoup(requests.get(url+"directory/"+j.lower()).content, 'html.parser')
            tables = soup.find_all('table')
            for k in tables:
                a = k.find_all('a')
                for l in a:
                    result = re.search('>(.*)<', str(l))
                    sub_subjects.append(result.group(1))
        sheet2.cell(count, column=1).value = i.value
        sheet2.cell(count, column=2).value = str(sub_subjects)
        count +=1
    wb.save('UTSHANDBOOKtest.xlsx')

def getAllSubjects(): # there were 32000 different subjects which 30931 were duplicate and 1038 subjects were unique (saving over 32x the space)
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOK.xlsx')
    subjects = []
    sheet1 = wb['Core Subjects']
    sheet2 = wb['SubMajor Subjects']
    sheet3 = wb['Major Subjects']
    sheet4 = wb['Subject Info']
    sheet4.cell(1, column=1).value = "Subject Code"
    CS = sheet1['B']
    MS = sheet2['B']
    SMS = sheet3['B']
    count = 2
    for i in CS:
        subject_codes = [x for x in re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value))) if x]
        subjects += subject_codes
    for i in MS:
        subject_codes = [x for x in re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value))) if x]
        subjects += subject_codes
    for i in SMS:
        subject_codes = [x for x in re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value))) if x]
        subjects += subject_codes    
    for i in subjects:
        sheet4.cell(count, column=1).value = i
        count+=1
    wb.save('UTSHANDBOOK.xlsx')
    
def getSubjectInfo():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOK.xlsx')
    sheet1 = wb['Subject Info']
    subjects = sheet1['A']
    sheet1.cell(1, column=1).value = "Pre-Requisite(s)"
    count = 1
    for i in subjects:
        soup = BeautifulSoup(requests.get(url+"subjects/details/"+str(i.value)).content, 'html.parser')
        em = soup.find_all('em')
        for j in em:
            if "Requisite(s):" in str(j):
                requisite = re.findall(r'\d+', str(j))
                for k in requisite: 
                    if len(str(k)) < 5:
                        requisite.remove(k)
                requisite = list(set(requisite))
                sheet1.cell(count, column=2).value = str(requisite)
                print(requisite)
        count += 1    
    wb.save('UTSHANDBOOK.xlsx')


def createMajorInfo():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    majorcodes = ws['B']
    ws2 = wb['Major Info']
    count = 1
    for i in majorcodes:
        majors = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        majors = [x for x in values1 if x]
        for j in majors:
            soup = BeautifulSoup(requests.get(url+"/directory/"+j.lower()).content, 'html.parser')
            title = soup.find(class_='ie-images').find('h1')
            ws2.cell(count, column=1).value = j
            ws2.cell(count, column=2).value = title.text
            print(j)
            count = count + 1
    wb.save('UTSHANDBOOKtest.xlsx')

def createSubMajorInfo():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    Subcodes = ws['c']
    ws2 = wb['SubMajor Info']
    count = 1
    for i in Subcodes:
        submajors = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        majors = [x for x in values1 if x]
        for j in majors:
            soup = BeautifulSoup(requests.get(url+"/directory/"+j.lower()).content, 'html.parser')
            title = soup.find(class_='ie-images').find('h1')
            ws2.cell(count, column=1).value = j
            ws2.cell(count, column=2).value = title.text
            print(j)
            count = count + 1
    wb.save('UTSHANDBOOKtest.xlsx')

def createCoreInfo():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    corecodes = ws['d']
    ws2 = wb['Core Info']
    count = 1
    for i in corecodes:
        core = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        majors = [x for x in values1 if x]
        for j in majors:
            soup = BeautifulSoup(requests.get(url+"/directory/"+j.lower()).content, 'html.parser')
            title = soup.find(class_='ie-images').find('h1')
            ws2.cell(count, column=1).value = j
            ws2.cell(count, column=2).value = title.text
            print(j)
            count = count + 1
    wb.save('UTSHANDBOOKtest.xlsx')

def getSubjectInfo2():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest2.xlsx')
    sheet1 = wb['Subject Info']
    subjects = sheet1['A']
    sheet1.cell(1, column=2).value = "Description"
    sheet1.cell(1, column=3).value = "SLO"
    sheet1.cell(1, column=4).value = "CILO"
    sheet1.cell(1, column=5).value = "Minimum Requirements"
    sheet1.cell(1, column=6).value = "Name"
    count = 1
    for i in subjects:
        soup = BeautifulSoup(requests.get(url+"subjects/details/"+str(i.value)).content, 'html.parser')
        SLO = []
        CILO = []
        for header in soup.find_all('h3'): #description
            if(header.text == "Description"):
                sheet1.cell(count, column=2).value = header.findNext('p').contents[0].text
        try:
            slotable = soup.find("table",{"class":"SLOTable"})
            for j in slotable.find_all('td'):
                SLO.append(j.text)
            sheet1.cell(count, column=3).value = str(SLO)
            SLO = []
        except Exception as e: 
            print(e)
        try:
            cilolist = soup.find("ul",{"class":"CILOList"})
            for k in cilolist.find_all('li'):
                CILO.append(k.text)
            sheet1.cell(count, column=4).value = str(CILO)
            CILO = []
        except:
            print("CILO")
        for header in soup.find_all('h3'): #description
            if(header.text == "Minimum requirements"):
                sheet1.cell(count, column=5).value = header.findNext('p').contents[0].text
        sheet1.cell(count, column=6).value = soup.find("h1").text
        print(soup.find("h1").text)
        count += 1    
    wb.save('UTSHANDBOOKtest2.xlsx')

def test():
    soup = BeautifulSoup(requests.get(url+"subjects/details/22108.html").content, 'html.parser')
    SLO = []
    CILO = []
    #print(soup)
    # for header in soup.find_all('h3'): #description
    #     if(header.text == "Description"):
    #         print(header.findNext('p').contents[0])
    try:
        slotable = soup.find("table",{"class":"SLOTable"})
        for j in slotable.find_all('td'):
            SLO.append(j.text)
        print(SLO)
        SLO = []
    except:
        print("works")
    # try:
    #     cilolist = soup.find("ul",{"class":"CILOList"})
    #     for j in cilolist.find_all('li'):
    #         CILO.append(j.text)
    #     print(CILO)
    # except:
    #     print("no cilo") 

def getMajorSubjectsRelo():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest3.xlsx')
    sheet1 = wb['Degree MAJ, SUB, CORE']
    sheet2 = wb['Sub Sub Relo']
    degrees = sheet1['D']
    sheet2.cell(1, column=1).value = "Degree Code"
    sheet2.cell(1, column=2).value = "Major Subjects"
    count = 2
    for i in degrees:
        core_subjects = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        major_codes = [x for x in values1 if x]
        print(major_codes)
        for j in major_codes:
            soup = BeautifulSoup(requests.get(url+"directory/"+j.lower()).content, 'html.parser')
            tables = soup.find_all('table')
            for k in tables:
                a = k.find_all('a')
                for l in a:
                    result = re.search('>(.*)<', str(l))
                    sheet2.cell(count, column=1).value = j
                    sheet2.cell(count, column=2).value = result.group(1)  
                    count +=1    
    wb.save('UTSHANDBOOKtest3.xlsx')

if __name__ == '__main__':
    #getMajors()
    #getDegrees()
    #getCoreSubjects()
    #getMajorSubjects()
    #getSubMajorSubjects()
    #getAllSubjects()
    #getSubjectInfo()
    #createMajorInfo()
    #createSubMajorInfo()
    #createCoreInfo()
    #getSubjectInfo2()
    #test()
    #getMajorSubjectsRelo()
    