import requests
import openpyxl
from bs4 import BeautifulSoup
import re

url = "https://www.handbook.uts.edu.au/"

def createDegreeData(): 
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degrees']
    f = open("Database\degreedata.cql", "w")
    for row in ws.iter_rows():
        f.write("CREATE" + " (" + row[1].value + ":" + "Degree {name: '" + row[2].value + "', code:'" + row[1].value + "', faculty: '" + row[0].value + "'})\n" )

def createMajorData():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Major Info']
    f = open("Database\majordata.cql", "w")
    for row in ws.iter_rows():
        f.write("CREATE" + " (" + row[0].value + ":" + "Major {name: '" + row[1].value + "', code:'" + row[0].value + "'})\n" )

def createSubMajorData():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['SubMajor Info']
    f = open("Database\submajordata.cql", "w")
    for row in ws.iter_rows():
        f.write("CREATE" + " (" + row[0].value + ":" + "SubMajor {name: '" + row[1].value + "', code:'" + row[0].value + "'})\n" )

def createCoreData():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Core Info']
    f = open("Database\coredata.cql", "w")
    for row in ws.iter_rows():
        f.write("CREATE" + " (" + row[0].value + ":" + "Core {name: '" + row[1].value + "', code:'" + row[0].value + "'})\n" )

def createMajorRelationships():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    majorcodes = ws['B']
    ws2 = wb['Degrees']
    degrees = ws2['B']
    ws3 = wb['Major Relationships']

    count = 1
    for i, k in zip(majorcodes, degrees):
        majors = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        majors = [x for x in values1 if x]
        for j in majors:
            ws3.cell(count, column=1).value = k.value
            ws3.cell(count, column=2).value = j
            count= count +1
    wb.save('UTSHANDBOOKtest.xlsx')

def createSubMajorRelationships():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    submajorcodes = ws['C']
    ws2 = wb['Degrees']
    degrees = ws2['B']
    ws3 = wb['SubMajor Relationships']

    count = 1
    for i, k in zip(submajorcodes, degrees):
        majors = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        majors = [x for x in values1 if x]
        for j in majors:
            ws3.cell(count, column=1).value = k.value
            ws3.cell(count, column=2).value = j
            count= count +1
    wb.save('UTSHANDBOOKtest.xlsx')

def createCoreRelationships():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Degree MAJ, SUB, CORE']
    corecodes = ws['D']
    ws2 = wb['Degrees']
    degrees = ws2['B']
    ws3 = wb['Core Relationships']

    count = 1
    for i, k in zip(corecodes, degrees):
        core = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(i.value)))
        core = [x for x in values1 if x]
        for j in core:
            ws3.cell(count, column=1).value = k.value
            ws3.cell(count, column=2).value = j
            count= count +1
    wb.save('UTSHANDBOOKtest.xlsx')

def createSubjectRelationships():
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest.xlsx')
    ws = wb['Subject Info']
    subjectcodes = ws['A']
    prerequisites = ws['B']
    ws3 = wb['Subject Relationships']
    count = 1
    for i, k in zip(subjectcodes, prerequisites):
        prerequisites = []
        values1 =  re.split(r'\s', re.sub(r'[^\w]', ' ', str(k.value)))
        prerequisites = [x for x in values1 if x]
        for j in prerequisites:
            ws3.cell(count, column=1).value = i.value
            ws3.cell(count, column=2).value = j
            count= count +1
    wb.save('UTSHANDBOOKtest.xlsx')

def createSubjectData():  #"', SLO:'" + str(row[2].value) + "', CILO:'" + str(row[3].value)
    wb = openpyxl.load_workbook(r'c:UTSHANDBOOKtest2.xlsx')
    ws = wb['Subject Info']
    f = open("Database\subjectdata.cql", "w")
    count = 0
    for row in ws.iter_rows():
        f.write("CREATE" + " (" + "n" + str(count) + ":" + "Subject {name: '" + str(row[5].value) + "', code:'" + str(row[0].value) +  "', description: '"  + str(row[1].value)  +  "', passreq:'" + str(row[4].value) + "'  })\n" )
        count = count + 1

# Code to mass import due to time complexity of creating each line individually
# LOAD CSV WITH HEADERS FROM "file:///majorrelo.csv" as data
# MATCH(degree:Degree), (major:Major)
# WHERE degree.code = data.Degree and major.code = data.Major 
# CREATE (degree)-[:MAJOR]->(major)
# CREATE (degree)<-[:MAJOR]- (major)

# LOAD CSV WITH HEADERS FROM "file:///submajorrelo.csv" as data
# MATCH(degree:Degree), (submajor:SubMajor)
# WHERE degree.code = data.Degree and submajor.code = data.SubMajor
# CREATE (degree)-[:MAJOR]->(submajor)

# to remove all relo dupes match ()-[r]->() delete r;

# To return node and it's relationships
# MATCH (n:Degree {code: 'C10026'})-[r]-(b)
# RETURN r, n, b

# MATCH (n:Degree {code: 'C10026'})-[r]-(b)-[r2]-(b2:Subject)
# RETURN r, n, b, r2, b2



if __name__ == '__main__':
    #createDegreeData()
    #createMajorData()
    #createMajorRelationships()
    #createCoreData()
    #createSubMajorData()
    #createSubMajorRelationships()
    #createCoreRelationships()
    #createSubjectRelationships()
    #createSubjectData()

